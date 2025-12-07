from flask import Flask, request, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = "appointments.xlsx"

# Serve open.html when accessing the root URL
@app.route("/")
def index():
    # Make sure open.html is in the project directory
    return send_from_directory(".", "open.html")

# Serve other static files (like CSS, JS)
@app.route("/<path:filename>")
def static_files(filename):
    return send_from_directory(".", filename)

# Handle the form submission
@app.route("/submit", methods=["POST"])
def submit():
    data = request.json
    patient_name = data.get("patientName")
    email = data.get("email")
    phone_number = data.get("phoneNumber")
    gender = data.get("gender")
    category = data.get("category")
    doctor = data.get("doctor")

    try:
        # Check if the Excel file exists
        if os.path.exists(EXCEL_FILE):
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Patient Name", "Email", "Phone Number", "Gender", "Category", "Doctor"])

        # Add data to the Excel file
        sheet.append([patient_name, email, phone_number, gender, category, doctor])
        workbook.save(EXCEL_FILE)

        return jsonify({"message": "Booking saved successfully!"}), 200

    except Exception as e:
        print("Error while saving to Excel:", e)
        return jsonify({"message": "An error occurred while saving the booking."}), 500

if __name__ == "__main__":
    app.run(debug=True)



@app.route("/get_bookings", methods=["GET"])
def get_bookings():
    try:
        if os.path.exists(EXCEL_FILE):
            # Load the existing workbook
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
            
            # Extract bookings data (assuming columns: Patient Name, Email, Phone, Gender, Category, Doctor)
            bookings = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                bookings.append({
                    "patient_name": row[0],  # Column 1: Patient Name
                    "email": row[1],         # Column 2: Email
                    "phone_number": row[2],  # Column 3: Phone Number
                    "gender": row[3],        # Column 4: Gender
                    "category": row[4],      # Column 5: Category
                    "doctor": row[5]         # Column 6: Doctor Name
                })
            
            return jsonify(bookings), 200
        else:
            return jsonify([]), 200  # No bookings if file does not exist

    except Exception as e:
        print("Error while fetching bookings:", e)
        return jsonify({"message": "An error occurred while fetching bookings."}), 500
